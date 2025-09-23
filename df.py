import pandas as pd
from openpyxl import Workbook

# шлях до твого файлу
file_path = "Актуалізація нерезидентів.xlsx"

# читаємо всі аркуші
excel_data = pd.ExcelFile(file_path)
sheets_data = {sheet: excel_data.parse(sheet) for sheet in excel_data.sheet_names}

# --- Smoke Checklist ---
smoke_wb = Workbook()
smoke_ws = smoke_wb.active
smoke_ws.title = "Smoke Checklist"
smoke_ws.append(["Module", "Title", "Expected Result"])

for sheet, df in sheets_data.items():
    if "Title" in df.columns and "Expected Result Android" in df.columns:
        subset = df[["Title", "Expected Result Android"]].dropna().head(2)
        for _, row in subset.iterrows():
            smoke_ws.append([sheet, str(row["Title"]), str(row["Expected Result Android"])])

smoke_wb.save("Smoke_Checklist.xlsx")


# --- Regression Checklist ---
regression_wb = Workbook()
for sheet, df in sheets_data.items():
    ws = regression_wb.create_sheet(title=sheet[:30])
    ws.append(df.columns.tolist())
    for _, row in df.iterrows():
        ws.append([str(x) if pd.notna(x) else "" for x in row.tolist()])

# видаляємо дефолтний аркуш
if "Sheet" in regression_wb.sheetnames:
    del regression_wb["Sheet"]

regression_wb.save("Regression_Checklist.xlsx")

print("Готово: Smoke_Checklist.xlsx і Regression_Checklist.xlsx")
